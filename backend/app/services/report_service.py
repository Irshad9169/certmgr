"""Report generation: CSV, XLSX, PDF, JSON for inventory/expiry/compliance and
deployment & renewal history."""

from __future__ import annotations

import csv
import io
import json
from datetime import UTC, datetime
from typing import Any

from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.models.certificate import Certificate
from app.models.job import JobExecution
from app.services.audit_service import query_audit

logger = get_logger(__name__)


def _cert_rows(db: Session, certificate_ids: list[int] | None = None) -> list[dict[str, Any]]:
    q = db.query(Certificate)
    if certificate_ids:
        q = q.filter(Certificate.id.in_(certificate_ids))
    rows = []
    for c in q.all():
        rows.append({
            "id": c.id, "domain": c.domain, "sans": ",".join(c.sans or []),
            "issuer": c.issuer or "", "environment": c.environment,
            "status": c.status, "key_type": c.key_type, "key_size": c.key_size or "",
            "signature_algorithm": c.signature_algorithm or "",
            "created_at": c.created_at.isoformat() if c.created_at else "",
            "expires_at": c.valid_until.isoformat() if c.valid_until else "",
            "days_remaining": c.days_remaining if c.days_remaining is not None else "",
            "auto_renew": c.auto_renew, "renewal_status": c.renewal_status,
            "provider": c.provider_name, "imported": c.imported, "tags": ",".join(t.name for t in c.tags),
        })
    return rows


def _history_rows(db: Session, job_type: str | None = None) -> list[dict[str, Any]]:
    q = db.query(JobExecution)
    if job_type:
        q = q.filter(JobExecution.job_type == job_type)
    rows = []
    for e in q.order_by(JobExecution.created_at.desc()).limit(5000).all():
        rows.append({
            "id": e.id, "job_type": e.job_type, "certificate_id": e.certificate_id,
            "server_id": e.server_id, "status": e.status, "trigger": e.trigger,
            "exit_code": e.exit_code, "duration_ms": e.execution_time_ms,
            "started_at": e.started_at.isoformat() if e.started_at else "",
            "finished_at": e.finished_at.isoformat() if e.finished_at else "",
            "error": (e.error_message or "")[:500],
        })
    return rows


def generate_report(db: Session, report_type: str, fmt: str,
                    certificate_ids: list[int] | None = None,
                    date_from: datetime | None = None, date_to: datetime | None = None,
                    ) -> tuple[bytes, str]:
    """Returns (bytes, filename). fmt ∈ csv|xlsx|pdf|json.

    date_from/date_to currently only narrow the "audit" report (the only
    one backed by a query that supports it — see query_audit()).
    """
    # PDF has real page-width constraints CSV/XLSX/JSON don't; a report
    # with a lot of columns (inventory) needs a trimmed column set there
    # specifically, or the table overflows the page regardless of font
    # size / wrapping. None set here means "same as `headers`".
    pdf_headers: list[str] | None = None

    if report_type == "inventory":
        data = _cert_rows(db, certificate_ids)
        headers = ["id", "domain", "sans", "issuer", "environment", "status", "key_type",
                   "key_size", "signature_algorithm", "created_at", "expires_at",
                   "days_remaining", "auto_renew", "renewal_status", "provider", "imported", "tags"]
        pdf_headers = ["domain", "issuer", "environment", "status", "expires_at",
                       "days_remaining", "auto_renew", "key_type"]
        title = "Certificate Inventory"
    elif report_type == "expiry":
        data = sorted(_cert_rows(db, certificate_ids), key=lambda r: r["expires_at"] or "")
        headers = ["domain", "issuer", "expires_at", "days_remaining", "status", "auto_renew", "renewal_status"]
        data = [{k: r[k] for k in headers} for r in data]
        title = "Certificate Expiry Report"
    elif report_type == "renewal_history":
        data = _history_rows(db, "renew")
        headers = ["id", "certificate_id", "status", "trigger", "exit_code", "duration_ms",
                   "started_at", "finished_at", "error"]
        title = "Renewal History"
    elif report_type == "deployment_history":
        data = _history_rows(db, "deploy")
        headers = ["id", "certificate_id", "server_id", "status", "trigger", "exit_code",
                   "duration_ms", "started_at", "finished_at", "error"]
        title = "Deployment History"
    elif report_type == "failures":
        data = _history_rows(db)
        data = [r for r in data if r["status"] == "failed"]
        headers = ["id", "job_type", "certificate_id", "status", "trigger", "error", "started_at"]
        data = [{k: r[k] for k in headers} for r in data]
        title = "Failure Report"
    elif report_type == "audit":
        rows, _ = query_audit(db, date_from=date_from, date_to=date_to, limit=5000)
        data = [{"id": a.id, "username": a.username or "", "action": a.action,
                 "resource_type": a.resource_type or "", "resource_id": a.resource_id or "",
                 "result": a.result, "ip_address": a.ip_address or "",
                 "created_at": a.created_at.isoformat() if a.created_at else ""} for a in rows]
        headers = ["id", "username", "action", "resource_type", "resource_id", "result", "ip_address", "created_at"]
        title = "Audit Log"
        if date_from or date_to:
            title += f" ({date_from.date() if date_from else '…'} to {date_to.date() if date_to else '…'})"
    else:
        raise ValueError(f"Unknown report type: {report_type}")

    if fmt == "json":
        return json.dumps({"title": title, "generated_at": datetime.now(UTC).isoformat(), "data": data},
                          indent=2).encode(), f"{report_type}.json"
    if fmt == "csv":
        return _to_csv(headers, data), f"{report_type}.csv"
    if fmt == "xlsx":
        return _to_xlsx(title, headers, data), f"{report_type}.xlsx"
    if fmt == "pdf":
        return _to_pdf(title, pdf_headers or headers, data), f"{report_type}.pdf"
    raise ValueError(f"Unsupported format: {fmt}")


def _to_csv(headers: list[str], data: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers)
    writer.writeheader()
    for row in data:
        writer.writerow({k: row.get(k, "") for k in headers})
    return buf.getvalue().encode("utf-8-sig")


def _to_xlsx(title: str, headers: list[str], data: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]
    ws.append(headers)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in data:
        ws.append([row.get(h, "") for h in headers])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            12, min(40, max(len(str(headers[col_idx-1])), 8))
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Columns that regularly hold long free-text (a full issuer DN, an error
# message, a comma-joined SAN/tag list) get proportionally more of the
# page width; everything else splits what's left evenly.
_PDF_WIDE_COLUMNS = {"issuer", "error", "sans", "tags", "domain", "action"}


def _to_pdf(title: str, headers: list[str], data: list[dict]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    margin = 24
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title,
                            leftMargin=margin, rightMargin=margin, topMargin=margin, bottomMargin=margin)
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle("ReportHeader", parent=styles["Normal"], fontName="Helvetica-Bold",
                                  fontSize=8, leading=10, textColor=colors.white, alignment=TA_LEFT)
    # wordWrap="CJK" also breaks long unbroken strings (hashes, tokens,
    # comma-free DNs) mid-word — without it, reportlab only wraps at
    # spaces, so exactly the kind of long unbroken value that overflows a
    # fixed-width column wouldn't wrap at all.
    cell_style = ParagraphStyle("ReportCell", parent=styles["Normal"], fontName="Helvetica",
                                fontSize=7, leading=9, alignment=TA_LEFT, wordWrap="CJK")

    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    weights = [3.0 if h in _PDF_WIDE_COLUMNS else 1.0 for h in headers]
    available_width = landscape(A4)[0] - 2 * margin
    col_widths = [available_width * w / sum(weights) for w in weights]

    table_data = [[Paragraph(str(h), header_style) for h in headers]]
    table_data += [[Paragraph(str(row.get(h, "")), cell_style) for h in headers] for row in data[:2000]]
    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()
